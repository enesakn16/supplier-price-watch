from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from enum import Enum
from pathlib import Path
from types import MappingProxyType
from typing import Iterable, Mapping

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MONEY_QUANTUM = Decimal("0.01")
PERCENT_QUANTUM = Decimal("0.01")
ZERO = Decimal("0")
HUNDRED = Decimal("100")
REQUIRED_CSV_COLUMNS = frozenset({"supplier", "sku", "unit_cost"})
CANONICAL_CSV_COLUMNS = REQUIRED_CSV_COLUMNS | {"currency"}


class PriceWatchError(ValueError):
    """Raised when supplier-price input is incomplete or financially unsafe."""


class RiskLevel(str, Enum):
    OK = "ok"
    WARNING = "warning"
    CRITICAL = "critical"


def _decimal(value: object, *, field: str) -> Decimal:
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise PriceWatchError(f"{field} must be a valid decimal value") from exc
    if not result.is_finite():
        raise PriceWatchError(f"{field} must be finite")
    return result


def _money(value: object, *, field: str) -> Decimal:
    result = _decimal(value, field=field)
    if result < ZERO:
        raise PriceWatchError(f"{field} cannot be negative")
    return result.quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)


@dataclass(frozen=True, slots=True)
class SupplierQuote:
    supplier: str
    sku: str
    unit_cost: Decimal
    currency: str = "TRY"

    @classmethod
    def from_mapping(cls, row: Mapping[str, object]) -> "SupplierQuote":
        supplier = str(row.get("supplier", "")).strip()
        sku = str(row.get("sku", "")).strip()
        raw_currency = row.get("currency")
        currency = "TRY" if raw_currency is None else str(raw_currency).strip().upper() or "TRY"
        if not supplier:
            raise PriceWatchError("supplier is required")
        if not sku:
            raise PriceWatchError("sku is required")
        if len(currency) != 3 or not currency.isalpha():
            raise PriceWatchError("currency must be a 3-letter code")
        return cls(
            supplier=supplier,
            sku=sku,
            unit_cost=_money(row.get("unit_cost"), field="unit_cost"),
            currency=currency,
        )


@dataclass(frozen=True, slots=True)
class SupplierImportProfile:
    """Explicit, versioned schema contract for one supplier price-list format."""

    profile_id: str
    supplier: str
    version: int
    column_map: Mapping[str, str]
    sheet_name: str | None = None

    def __post_init__(self) -> None:
        profile_id = self.profile_id.strip()
        supplier = self.supplier.strip()
        if not profile_id:
            raise PriceWatchError("profile_id is required")
        if not supplier:
            raise PriceWatchError("profile supplier is required")
        if self.version < 1:
            raise PriceWatchError("profile version must be at least 1")
        normalized = _normalize_column_map(self.column_map)
        object.__setattr__(self, "profile_id", profile_id)
        object.__setattr__(self, "supplier", supplier)
        object.__setattr__(self, "column_map", MappingProxyType(normalized))
        if self.sheet_name is not None:
            sheet_name = self.sheet_name.strip()
            if not sheet_name:
                raise PriceWatchError("profile sheet_name cannot be blank")
            object.__setattr__(self, "sheet_name", sheet_name)


class SupplierProfileRegistry:
    """Fail-closed registry that resolves exact profile IDs and versions only."""

    def __init__(self, profiles: Iterable[SupplierImportProfile] = ()) -> None:
        self._profiles: dict[tuple[str, int], SupplierImportProfile] = {}
        for profile in profiles:
            self.register(profile)

    def register(self, profile: SupplierImportProfile) -> None:
        key = (profile.profile_id, profile.version)
        if key in self._profiles:
            raise PriceWatchError(
                f"duplicate supplier profile/version: {profile.profile_id} v{profile.version}"
            )
        self._profiles[key] = profile

    def get(self, profile_id: str, *, version: int | None = None) -> SupplierImportProfile:
        normalized_id = profile_id.strip()
        if not normalized_id:
            raise PriceWatchError("profile_id is required")

        if version is not None:
            profile = self._profiles.get((normalized_id, version))
            if profile is None:
                raise PriceWatchError(f"supplier profile not found: {normalized_id} v{version}")
            return profile

        matches = [
            profile
            for (candidate_id, _), profile in self._profiles.items()
            if candidate_id == normalized_id
        ]
        if not matches:
            raise PriceWatchError(f"supplier profile not found: {normalized_id}")
        return max(matches, key=lambda profile: profile.version)

    def list_profiles(self) -> tuple[SupplierImportProfile, ...]:
        return tuple(
            sorted(
                self._profiles.values(),
                key=lambda profile: (profile.profile_id, profile.version),
            )
        )


def _normalize_column_map(column_map: Mapping[str, str] | None) -> dict[str, str]:
    if column_map is None:
        return {}

    normalized: dict[str, str] = {}
    targets: set[str] = set()
    for raw_source, raw_target in column_map.items():
        source = str(raw_source).strip()
        target = str(raw_target).strip()
        if not source or not target:
            raise PriceWatchError("column map names cannot be empty")
        if source in normalized:
            raise PriceWatchError(f"column map contains duplicate source: {source}")
        if target not in CANONICAL_CSV_COLUMNS:
            raise PriceWatchError(f"column map target is not supported: {target}")
        if target in targets:
            raise PriceWatchError(f"column map maps multiple columns to: {target}")
        normalized[source] = target
        targets.add(target)
    return normalized


def _apply_supplier_profile(
    quotes: list[SupplierQuote],
    profile: SupplierImportProfile,
) -> list[SupplierQuote]:
    """Bind a profile's trusted supplier identity to parsed rows.

    Supplier names embedded in source files are treated as data, not authority. If
    the file includes a non-empty supplier value it must agree with the profile.
    Missing supplier values are safely populated from the selected profile.
    """

    bound: list[SupplierQuote] = []
    for quote in quotes:
        if quote.supplier != profile.supplier:
            raise PriceWatchError(
                f"supplier profile mismatch: expected {profile.supplier}, got {quote.supplier}"
            )
        bound.append(quote)
    return bound


def load_quotes_csv(
    path: str | Path,
    *,
    column_map: Mapping[str, str] | None = None,
) -> list[SupplierQuote]:
    """Load a supplier quote CSV with strict schema and duplicate validation.

    Required canonical columns are ``supplier``, ``sku`` and ``unit_cost``.
    ``currency`` is optional and defaults to TRY. Supplier-specific headers can be
    mapped explicitly with ``column_map={"source header": "canonical_name"}``.
    Header guessing is intentionally forbidden.
    """

    csv_path = Path(path)
    mapping = _normalize_column_map(column_map)

    try:
        handle = csv_path.open("r", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise PriceWatchError(f"cannot read CSV: {csv_path}") from exc

    with handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise PriceWatchError("CSV header is required")

        headers = [name.strip() if name is not None else "" for name in reader.fieldnames]
        if any(not name for name in headers):
            raise PriceWatchError("CSV contains an empty column name")
        if len(headers) != len(set(headers)):
            raise PriceWatchError("CSV contains duplicate column names")

        missing_sources = sorted(set(mapping).difference(headers))
        if missing_sources:
            raise PriceWatchError(
                f"CSV column map references missing columns: {', '.join(missing_sources)}"
            )

        canonical_headers = [mapping.get(name, name) for name in headers]
        if len(canonical_headers) != len(set(canonical_headers)):
            raise PriceWatchError("CSV column map creates duplicate canonical columns")

        missing = sorted(REQUIRED_CSV_COLUMNS.difference(canonical_headers))
        if missing:
            raise PriceWatchError(f"CSV missing required columns: {', '.join(missing)}")

        quotes: list[SupplierQuote] = []
        seen: set[tuple[str, str, str]] = set()
        for line_number, raw_row in enumerate(reader, start=2):
            if None in raw_row:
                raise PriceWatchError(f"CSV row {line_number} has more values than headers")

            row = {
                mapping.get(str(key).strip(), str(key).strip()): value
                for key, value in raw_row.items()
            }
            if all(value is None or not str(value).strip() for value in row.values()):
                continue

            try:
                quote = SupplierQuote.from_mapping(row)
            except PriceWatchError as exc:
                raise PriceWatchError(f"CSV row {line_number}: {exc}") from exc

            identity = (quote.supplier, quote.sku, quote.currency)
            if identity in seen:
                raise PriceWatchError(
                    f"CSV row {line_number}: duplicate supplier/SKU/currency identity"
                )
            seen.add(identity)
            quotes.append(quote)

        return quotes


def load_quotes_csv_profile(
    path: str | Path,
    profile: SupplierImportProfile,
) -> list[SupplierQuote]:
    """Load CSV using one explicit supplier schema profile."""

    return _apply_supplier_profile(
        load_quotes_csv(path, column_map=profile.column_map),
        profile,
    )


def load_quotes_xlsx(
    path: str | Path,
    *,
    column_map: Mapping[str, str] | None = None,
    sheet_name: str | None = None,
) -> list[SupplierQuote]:
    """Load one XLSX worksheet using the same explicit schema rules as CSV.

    The workbook is opened in read-only/data-only mode. When ``sheet_name`` is
    omitted, the active worksheet is used. Formula cells are never evaluated by
    this loader; callers must provide workbooks with cached values or plain data.
    """

    xlsx_path = Path(path)
    mapping = _normalize_column_map(column_map)

    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except (OSError, InvalidFileException, ValueError) as exc:
        raise PriceWatchError(f"cannot read XLSX: {xlsx_path}") from exc

    try:
        if sheet_name is None:
            worksheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                raise PriceWatchError(f"XLSX worksheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]

        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise PriceWatchError("XLSX header is required") from exc

        headers = ["" if value is None else str(value).strip() for value in raw_headers]
        while headers and not headers[-1]:
            headers.pop()

        if not headers:
            raise PriceWatchError("XLSX header is required")
        if any(not name for name in headers):
            raise PriceWatchError("XLSX contains an empty column name")
        if len(headers) != len(set(headers)):
            raise PriceWatchError("XLSX contains duplicate column names")

        missing_sources = sorted(set(mapping).difference(headers))
        if missing_sources:
            raise PriceWatchError(
                f"XLSX column map references missing columns: {', '.join(missing_sources)}"
            )

        canonical_headers = [mapping.get(name, name) for name in headers]
        if len(canonical_headers) != len(set(canonical_headers)):
            raise PriceWatchError("XLSX column map creates duplicate canonical columns")

        missing = sorted(REQUIRED_CSV_COLUMNS.difference(canonical_headers))
        if missing:
            raise PriceWatchError(f"XLSX missing required columns: {', '.join(missing)}")

        quotes: list[SupplierQuote] = []
        seen: set[tuple[str, str, str]] = set()
        width = len(headers)

        for row_number, values in enumerate(rows, start=2):
            materialized = list(values)
            extra_values = materialized[width:]
            if any(value is not None and str(value).strip() for value in extra_values):
                raise PriceWatchError(f"XLSX row {row_number} has more values than headers")

            materialized = materialized[:width]
            if len(materialized) < width:
                materialized.extend([None] * (width - len(materialized)))

            if all(value is None or not str(value).strip() for value in materialized):
                continue

            row = {
                mapping.get(header, header): value
                for header, value in zip(headers, materialized, strict=True)
            }
            try:
                quote = SupplierQuote.from_mapping(row)
            except PriceWatchError as exc:
                raise PriceWatchError(f"XLSX row {row_number}: {exc}") from exc

            identity = (quote.supplier, quote.sku, quote.currency)
            if identity in seen:
                raise PriceWatchError(
                    f"XLSX row {row_number}: duplicate supplier/SKU/currency identity"
                )
            seen.add(identity)
            quotes.append(quote)

        return quotes
    finally:
        workbook.close()


def load_quotes_xlsx_profile(
    path: str | Path,
    profile: SupplierImportProfile,
) -> list[SupplierQuote]:
    """Load XLSX using one explicit supplier schema profile."""

    return _apply_supplier_profile(
        load_quotes_xlsx(
            path,
            column_map=profile.column_map,
            sheet_name=profile.sheet_name,
        ),
        profile,
    )


@dataclass(frozen=True, slots=True)
class PriceComparison:
    supplier: str
    sku: str
    previous_cost: Decimal
    current_cost: Decimal
    absolute_change: Decimal
    percent_change: Decimal | None
    gross_margin_percent: Decimal | None
    risk: RiskLevel


def compare_quote(
    previous: SupplierQuote,
    current: SupplierQuote,
    *,
    sale_price: object | None = None,
    warning_margin_percent: object = "20",
    critical_margin_percent: object = "10",
) -> PriceComparison:
    """Compare one supplier quote with its prior value using Decimal math.

    The comparison fails closed when supplier, SKU or currency differ. Margin risk
    is only calculated when a sale price is explicitly provided.
    """

    if previous.supplier != current.supplier:
        raise PriceWatchError("supplier mismatch")
    if previous.sku != current.sku:
        raise PriceWatchError("sku mismatch")
    if previous.currency != current.currency:
        raise PriceWatchError("currency mismatch; convert explicitly before comparison")

    absolute_change = (current.unit_cost - previous.unit_cost).quantize(
        MONEY_QUANTUM, rounding=ROUND_HALF_UP
    )
    percent_change = None
    if previous.unit_cost != ZERO:
        percent_change = (
            absolute_change / previous.unit_cost * HUNDRED
        ).quantize(PERCENT_QUANTUM, rounding=ROUND_HALF_UP)

    gross_margin_percent = None
    risk = RiskLevel.OK

    if sale_price is not None:
        sale = _money(sale_price, field="sale_price")
        if sale == ZERO:
            raise PriceWatchError("sale_price must be greater than zero")

        warning = _decimal(warning_margin_percent, field="warning_margin_percent")
        critical = _decimal(critical_margin_percent, field="critical_margin_percent")
        if warning < ZERO or critical < ZERO:
            raise PriceWatchError("margin thresholds cannot be negative")
        if critical > warning:
            raise PriceWatchError("critical margin threshold cannot exceed warning threshold")

        gross_margin_percent = (
            (sale - current.unit_cost) / sale * HUNDRED
        ).quantize(PERCENT_QUANTUM, rounding=ROUND_HALF_UP)

        if gross_margin_percent <= critical:
            risk = RiskLevel.CRITICAL
        elif gross_margin_percent <= warning:
            risk = RiskLevel.WARNING

    return PriceComparison(
        supplier=current.supplier,
        sku=current.sku,
        previous_cost=previous.unit_cost,
        current_cost=current.unit_cost,
        absolute_change=absolute_change,
        percent_change=percent_change,
        gross_margin_percent=gross_margin_percent,
        risk=risk,
    )


def compare_catalogs(
    previous_quotes: Iterable[SupplierQuote],
    current_quotes: Iterable[SupplierQuote],
    *,
    sale_prices: Mapping[str, object] | None = None,
) -> list[PriceComparison]:
    """Compare matching supplier/SKU/currency rows and ignore unmatched rows.

    Unmatched items are intentionally not guessed. Callers can separately report
    additions/removals instead of accidentally pairing unrelated supplier data.
    """

    old_index = {
        (quote.supplier, quote.sku, quote.currency): quote for quote in previous_quotes
    }
    results: list[PriceComparison] = []

    for current in current_quotes:
        previous = old_index.get((current.supplier, current.sku, current.currency))
        if previous is None:
            continue
        sale_price = sale_prices.get(current.sku) if sale_prices else None
        results.append(compare_quote(previous, current, sale_price=sale_price))

    return results
