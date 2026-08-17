from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from supplier_price_watch import (
    PriceWatchError,
    SupplierImportProfile,
    SupplierProfileRegistry,
    load_quotes_csv_profile,
    load_quotes_xlsx_profile,
)


class SupplierImportProfileTests(unittest.TestCase):
    def profile(self, *, version: int = 1, sheet_name: str | None = None):
        return SupplierImportProfile(
            profile_id="arzu-standard",
            supplier="Arzu",
            version=version,
            column_map={
                "Tedarikci": "supplier",
                "Urun Kodu": "sku",
                "Alis Fiyati": "unit_cost",
                "Para Birimi": "currency",
            },
            sheet_name=sheet_name,
        )

    def supplierless_profile(self, *, sheet_name: str | None = None):
        return SupplierImportProfile(
            profile_id="arzu-supplierless",
            supplier="Arzu",
            version=1,
            column_map={
                "Urun Kodu": "sku",
                "Alis Fiyati": "unit_cost",
                "Para Birimi": "currency",
            },
            sheet_name=sheet_name,
        )

    def test_profile_mapping_is_immutable(self):
        profile = self.profile()

        with self.assertRaises(TypeError):
            profile.column_map["Fiyat"] = "unit_cost"

    def test_profile_rejects_blank_identity_and_invalid_version(self):
        with self.assertRaisesRegex(PriceWatchError, "profile_id is required"):
            SupplierImportProfile("  ", "Arzu", 1, {})

        with self.assertRaisesRegex(PriceWatchError, "profile supplier is required"):
            SupplierImportProfile("arzu", "  ", 1, {})

        with self.assertRaisesRegex(PriceWatchError, "version must be at least 1"):
            SupplierImportProfile("arzu", "Arzu", 0, {})

    def test_registry_rejects_duplicate_profile_version(self):
        profile = self.profile()
        registry = SupplierProfileRegistry([profile])

        with self.assertRaisesRegex(PriceWatchError, "duplicate supplier profile/version"):
            registry.register(profile)

    def test_registry_resolves_latest_or_exact_version(self):
        v1 = self.profile(version=1)
        v2 = self.profile(version=2)
        registry = SupplierProfileRegistry([v1, v2])

        self.assertIs(registry.get("arzu-standard"), v2)
        self.assertIs(registry.get("arzu-standard", version=1), v1)

        with self.assertRaisesRegex(PriceWatchError, "not found"):
            registry.get("arzu-standard", version=3)

    def test_csv_profile_rejects_supplier_mismatch(self):
        temp_dir = TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "quotes.csv"
        path.write_text(
            "Tedarikci,Urun Kodu,Alis Fiyati,Para Birimi\n"
            "MSR,SKU-1,100,TRY\n",
            encoding="utf-8",
        )

        with self.assertRaisesRegex(PriceWatchError, "supplier profile mismatch"):
            load_quotes_csv_profile(path, self.profile())

    def test_csv_profile_loads_matching_supplier(self):
        temp_dir = TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "quotes.csv"
        path.write_text(
            "Tedarikci,Urun Kodu,Alis Fiyati,Para Birimi\n"
            "Arzu,SKU-1,100.25,try\n",
            encoding="utf-8",
        )

        quotes = load_quotes_csv_profile(path, self.profile())

        self.assertEqual(len(quotes), 1)
        self.assertEqual(quotes[0].supplier, "Arzu")
        self.assertEqual(quotes[0].sku, "SKU-1")
        self.assertEqual(quotes[0].currency, "TRY")

    def test_csv_profile_can_supply_missing_supplier_column(self):
        temp_dir = TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "quotes.csv"
        path.write_text(
            "Urun Kodu,Alis Fiyati,Para Birimi\n"
            "SKU-10,125.40,TRY\n",
            encoding="utf-8",
        )

        quotes = load_quotes_csv_profile(path, self.supplierless_profile())

        self.assertEqual(len(quotes), 1)
        self.assertEqual(quotes[0].supplier, "Arzu")
        self.assertEqual(quotes[0].sku, "SKU-10")
        self.assertEqual(quotes[0].unit_cost, 125.40)

    def test_xlsx_profile_uses_declared_sheet(self):
        temp_dir = TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "quotes.xlsx"

        workbook = Workbook()
        ignored = workbook.active
        ignored.title = "Ignore"
        ignored.append(["Tedarikci", "Urun Kodu", "Alis Fiyati", "Para Birimi"])
        ignored.append(["MSR", "WRONG", 1, "TRY"])

        sheet = workbook.create_sheet("Arzu Listesi")
        sheet.append(["Tedarikci", "Urun Kodu", "Alis Fiyati", "Para Birimi"])
        sheet.append(["Arzu", "SKU-2", 77.5, "TRY"])
        workbook.save(path)
        workbook.close()

        quotes = load_quotes_xlsx_profile(
            path,
            self.profile(sheet_name="Arzu Listesi"),
        )

        self.assertEqual(len(quotes), 1)
        self.assertEqual(quotes[0].supplier, "Arzu")
        self.assertEqual(quotes[0].sku, "SKU-2")

    def test_xlsx_profile_can_supply_missing_supplier_column(self):
        temp_dir = TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "quotes.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Arzu Listesi"
        sheet.append(["Urun Kodu", "Alis Fiyati", "Para Birimi"])
        sheet.append(["SKU-20", 88.75, "TRY"])
        workbook.save(path)
        workbook.close()

        quotes = load_quotes_xlsx_profile(
            path,
            self.supplierless_profile(sheet_name="Arzu Listesi"),
        )

        self.assertEqual(len(quotes), 1)
        self.assertEqual(quotes[0].supplier, "Arzu")
        self.assertEqual(quotes[0].sku, "SKU-20")
        self.assertEqual(quotes[0].unit_cost, 88.75)


if __name__ == "__main__":
    unittest.main()
